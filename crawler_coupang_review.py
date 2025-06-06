from bs4 import BeautifulSoup as bs
from openpyxl import Workbook
from fake_useragent import UserAgent
from requests.exceptions import RequestException, Timeout, ConnectTimeout, ReadTimeout
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
import time
import os
import re
import requests as rq
import random
import json


class NonWindowsUserAgent:
    """Windows를 제외한 User-Agent 생성기 (fake_useragent 기반)"""

    def __init__(self):
        self.ua = UserAgent()
        self.max_attempts = 20  # 최대 시도 횟수

        # Windows 관련 키워드들
        self.windows_keywords = [
            'Windows NT', 'Win32', 'Win64', 'WOW64', 'Windows 10', 'Windows 11',
            'Windows 7', 'Windows 8', 'Microsoft Windows', 'win32', 'win64'
        ]

    def _is_windows_ua(self, user_agent):
        """User-Agent가 Windows인지 확인"""
        if not user_agent:
            return True

        user_agent_lower = user_agent.lower()
        return any(keyword.lower() in user_agent_lower for keyword in self.windows_keywords)

    def _get_non_windows_ua(self, ua_type='random'):
        """Windows가 아닌 User-Agent를 가져오기"""
        for attempt in range(self.max_attempts):
            try:
                if ua_type == 'chrome':
                    user_agent = self.ua.chrome
                elif ua_type == 'firefox':
                    user_agent = self.ua.firefox
                elif ua_type == 'safari':
                    user_agent = self.ua.safari
                else:
                    user_agent = self.ua.random

                if not self._is_windows_ua(user_agent):
                    print(f"[DEBUG] User-Agent 선택 성공 (시도 {attempt + 1}회): {user_agent[:50]}...")
                    return user_agent
                else:
                    print(f"[DEBUG] Windows UA 감지, 재시도 중... (시도 {attempt + 1}/{self.max_attempts})")

            except Exception as e:
                print(f"[WARNING] User-Agent 생성 오류 (시도 {attempt + 1}): {e}")
                continue

        # 모든 시도가 실패한 경우 안전한 Mac UA 반환
        print(f"[WARNING] {self.max_attempts}회 시도 후에도 적절한 UA를 찾지 못했습니다. 기본 Mac UA 사용")
        return "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"

    @property
    def random(self):
        """랜덤한 Non-Windows User-Agent 반환"""
        return self._get_non_windows_ua('random')

    @property
    def chrome(self):
        """Chrome Non-Windows User-Agent 반환"""
        return self._get_non_windows_ua('chrome')

    @property
    def firefox(self):
        """Firefox Non-Windows User-Agent 반환"""
        return self._get_non_windows_ua('firefox')

    @property
    def safari(self):
        """Safari Non-Windows User-Agent 반환"""
        return self._get_non_windows_ua('safari')

    def get_mobile_ua(self):
        """모바일 전용 User-Agent 반환"""
        for attempt in range(self.max_attempts):
            try:
                # 모바일 브라우저 위주로 시도
                browser_types = ['chrome', 'safari', 'random']
                ua_type = random.choice(browser_types)

                user_agent = self._get_non_windows_ua(ua_type)

                # 모바일 키워드가 있는지 확인
                mobile_keywords = ['Mobile', 'Android', 'iPhone', 'iPad']
                if any(keyword in user_agent for keyword in mobile_keywords):
                    return user_agent

            except Exception as e:
                print(f"[WARNING] 모바일 UA 생성 오류: {e}")
                continue

        # 실패 시 안전한 Android UA 반환
        return "Mozilla/5.0 (Linux; Android 10; SM-G973F) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Mobile Safari/537.36"

    def get_desktop_ua(self):
        """데스크톱 전용 User-Agent 반환 (Mac 위주)"""
        for attempt in range(self.max_attempts):
            try:
                user_agent = self._get_non_windows_ua('random')

                # 데스크톱이면서 모바일이 아닌 것
                if 'Macintosh' in user_agent and 'Mobile' not in user_agent:
                    return user_agent

            except Exception as e:
                print(f"[WARNING] 데스크톱 UA 생성 오류: {e}")
                continue

        # 실패 시 안전한 Mac UA 반환
        return "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.1 Safari/605.1.15"


class ProxyRotator:
    def __init__(self, proxy_list=None):
        """
        프록시 로테이터 초기화
        proxy_list: ['ip:port:username:password', ...] 형태의 프록시 리스트
        """
        self.proxy_list = proxy_list if proxy_list else []
        # itertools.cycle 제거 - 랜덤 선택으로 변경
        self.current_proxy = None
        self.failed_proxies = set()
        self.proxy_failure_count = {}  # 프록시별 실패 횟수 추적
        self.max_failures_per_proxy = 3  # 프록시당 최대 실패 허용 횟수

    def get_next_proxy(self):
        """랜덤하게 프록시를 선택하여 반환"""
        if not self.proxy_list:
            return None

        # 사용 가능한 프록시 목록 생성
        available_proxies = [proxy for proxy in self.proxy_list if proxy not in self.failed_proxies]

        if not available_proxies:
            # 모든 프록시가 완전히 실패했다면 실패 목록을 초기화
            print("[WARNING] 모든 프록시가 실패했습니다. 실패 목록과 카운터를 초기화합니다.")
            self.failed_proxies.clear()
            self.proxy_failure_count.clear()
            available_proxies = self.proxy_list.copy()

        # 사용 가능한 프록시 중에서 랜덤하게 선택
        proxy = random.choice(available_proxies)
        self.current_proxy = proxy

        proxy_ip = proxy.split(':')[0]
        failure_count = self.proxy_failure_count.get(proxy, 0)
        print(f"[PROXY] 랜덤 선택된 프록시: {proxy_ip} (실패 횟수: {failure_count})")
        return proxy

    def get_random_proxy_from_working_set(self):
        """성능이 좋은 프록시들 중에서 랜덤 선택"""
        if not self.proxy_list:
            return None

        # 실패 횟수가 적은 프록시들을 우선적으로 선택
        working_proxies = []
        for proxy in self.proxy_list:
            if proxy not in self.failed_proxies:
                failure_count = self.proxy_failure_count.get(proxy, 0)
                # 실패 횟수가 1회 이하인 프록시들을 우선 선택
                if failure_count <= 1:
                    working_proxies.append(proxy)

        # 성능 좋은 프록시가 없으면 일반 선택 방식 사용
        if not working_proxies:
            return self.get_next_proxy()

        proxy = random.choice(working_proxies)
        self.current_proxy = proxy
        proxy_ip = proxy.split(':')[0]
        failure_count = self.proxy_failure_count.get(proxy, 0)
        print(f"[PROXY] 성능 우선 랜덤 선택: {proxy_ip} (실패 횟수: {failure_count})")
        return proxy

    def mark_proxy_failed(self, proxy):
        """프록시를 실패로 표시 (누적 실패 관리)"""
        if proxy not in self.proxy_failure_count:
            self.proxy_failure_count[proxy] = 0

        self.proxy_failure_count[proxy] += 1
        proxy_ip = proxy.split(':')[0]

        # 최대 실패 횟수에 도달하면 완전히 제거
        if self.proxy_failure_count[proxy] >= self.max_failures_per_proxy:
            self.failed_proxies.add(proxy)
            print(f"[WARNING] 프록시 완전 실패로 제거: {proxy_ip} ({self.proxy_failure_count[proxy]}회 실패)")
        else:
            print(
                f"[WARNING] 프록시 일시 실패: {proxy_ip} ({self.proxy_failure_count[proxy]}/{self.max_failures_per_proxy} 실패)")

    def get_available_proxy_count(self):
        """사용 가능한 프록시 개수 반환"""
        if not self.proxy_list:
            return 0
        return len(self.proxy_list) - len(self.failed_proxies)

    def get_proxy_dict(self, proxy_string):
        """프록시 문자열을 requests용 딕셔너리로 변환"""
        if not proxy_string:
            return None

        parts = proxy_string.split(':')
        if len(parts) == 2:  # ip:port
            ip, port = parts
            return {
                'http': f'http://{ip}:{port}',
                'https': f'http://{ip}:{port}'
            }
        elif len(parts) == 4:  # ip:port:username:password
            ip, port, username, password = parts
            return {
                'http': f'http://{username}:{password}@{ip}:{port}',
                'https': f'http://{username}:{password}@{ip}:{port}'
            }
        return None


class ChromeDriver:
    def __init__(self, proxy_rotator=None) -> None:
        self.proxy_rotator = proxy_rotator
        self.ua = NonWindowsUserAgent()  # Windows 제외 User-Agent 사용
        self.set_options()
        self.set_driver()

    def set_options(self) -> None:
        self.options = Options()
        self.options.add_argument("--headless")
        self.options.add_argument("lang=ko_KR")

        # 모바일/Mac 전용 User-Agent 사용
        user_agent = self.ua.random
        self.options.add_argument(f"user-agent={user_agent}")
        print(f"[DEBUG] 사용 중인 User-Agent: {user_agent}")

        # 더 많은 브라우저 옵션 추가로 탐지 방지
        self.options.add_argument("--log-level=3")
        self.options.add_argument("--no-sandbox")
        self.options.add_argument("--disable-dev-shm-usage")
        self.options.add_argument("--disable-blink-features=AutomationControlled")
        self.options.add_argument("--exclude-switches=enable-automation")
        self.options.add_argument("--disable-extensions")
        self.options.add_argument("--no-first-run")
        self.options.add_argument("--disable-default-apps")
        self.options.add_argument("--disable-infobars")
        self.options.add_experimental_option("detach", True)
        self.options.add_experimental_option("excludeSwitches", ["enable-automation"])
        self.options.add_experimental_option('useAutomationExtension', False)

        # 프록시 설정
        if self.proxy_rotator:
            proxy = self.proxy_rotator.get_next_proxy()
            if proxy:
                parts = proxy.split(':')
                if len(parts) >= 2:
                    ip, port = parts[0], parts[1]
                    self.options.add_argument(f'--proxy-server=http://{ip}:{port}')
                    print(f"[DEBUG] Selenium 프록시 설정: {ip}:{port}")

    def set_driver(self) -> None:
        self.driver = webdriver.Chrome(options=self.options)
        # WebDriver 탐지 방지
        self.driver.execute_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")

    def refresh_with_new_proxy(self):
        """새로운 프록시로 드라이버 재시작"""
        if hasattr(self, 'driver') and self.driver:
            self.driver.quit()
        self.set_options()
        self.set_driver()


class URLManager:
    """URL 관리 클래스 (JSON 지원)"""

    def __init__(self, file_path="data/홈플래닛_products_dedup_first.json"):
        self.file_path = file_path
        self.products = []  # URL과 상품명을 함께 저장
        self.current_index = 0

    def load_urls_from_json(self):
        """JSON 파일에서 URL 목록과 상품명 로드"""
        try:
            if not os.path.exists(self.file_path):
                self.create_sample_file()
                return False

            with open(self.file_path, 'r', encoding='utf-8') as f:
                data = json.load(f)

            self.products = []

            # JSON이 리스트인 경우
            if isinstance(data, list):
                for i, item in enumerate(data, 1):
                    if isinstance(item, dict) and 'product_url' in item:
                        url = item['product_url']
                        product_name = item.get('product_name', f'상품_{i}')

                        if "coupang.com" in url and "products/" in url:
                            self.products.append({
                                'url': url,
                                'name': product_name
                            })
                        else:
                            print(f"[WARNING] 잘못된 URL 형식 (항목 {i}): {url}")

            if self.products:
                print(f"[INFO] {len(self.products)}개의 유효한 상품을 로드했습니다.")
                return True
            else:
                print("[ERROR] 유효한 상품이 없습니다.")
                return False

        except json.JSONDecodeError as e:
            print(f"[ERROR] JSON 파일 파싱 오류: {e}")
            return False
        except Exception as e:
            print(f"[ERROR] JSON 파일 읽기 실패: {e}")
            return False

    def create_sample_file(self):
        """샘플 JSON 파일 생성"""
        sample_data = [
            {
                "product_url": "https://www.coupang.com/vp/products/7335597976?itemId=18741704367&vendorItemId=85873964906",
                "product_name": "샘플 상품 1"
            },
            {
                "product_url": "https://www.coupang.com/vp/products/1234567890?itemId=12345678901&vendorItemId=98765432109",
                "product_name": "샘플 상품 2"
            }
        ]

        try:
            with open(self.file_path, 'w', encoding='utf-8') as f:
                json.dump(sample_data, f, ensure_ascii=False, indent=2)
            print(f"[INFO] 샘플 JSON 파일이 생성되었습니다: {self.file_path}")
            print("[INFO] 파일을 편집하여 크롤링할 상품 정보를 입력한 후 다시 실행하세요.")
        except Exception as e:
            print(f"[ERROR] 샘플 파일 생성 실패: {e}")

    def get_next_product(self):
        """다음 상품 정보 반환 (URL과 상품명)"""
        if self.current_index < len(self.products):
            product = self.products[self.current_index]
            self.current_index += 1
            return product
        return None

    def get_remaining_count(self):
        """남은 상품 개수 반환"""
        return len(self.products) - self.current_index

    def get_current_progress(self):
        """현재 진행률 반환"""
        return self.current_index, len(self.products)


class Coupang:
    @staticmethod
    def get_product_code(url: str) -> str:
        prod_code: str = url.split("products/")[-1].split("?")[0]
        return prod_code

    @staticmethod
    def get_soup_object(resp: rq.Response) -> bs:
        return bs(resp.text, "html.parser")

    def __del__(self) -> None:
        if hasattr(self, 'ch') and self.ch.driver:
            self.ch.driver.quit()

    def __init__(self, proxy_list=None) -> None:
        # delay 관련 설정
        self.base_review_url: str = "https://www.coupang.com/vp/product/reviews"
        self.retries = 10  # 재시도 횟수 줄임
        self.delay_min = 1.0  # 최소 딜레이 증가
        self.delay_max = 2.0  # 최대 딜레이 증가
        self.page_delay_min = 0.0  # 페이지 간 최소 딜레이 증가
        self.page_delay_max = 0.0  # 페이지 간 최대 딜레이 증가
        self.max_pages = 150  # v1.6: 최대 페이지를 300으로 제한

        # 타임아웃 관련 설정
        self.consecutive_timeouts = 0
        self.max_consecutive_timeouts = 5  # 연속 타임아웃 허용 횟수 감소
        self.long_wait_min = 10  # 긴 대기 시간 줄임 (5분)
        self.long_wait_max = 15  # 긴 대기 시간 줄임 (7분)

        # 프록시 로테이터 초기화
        self.proxy_rotator = ProxyRotator(proxy_list)

        # Windows 제외 User-Agent 초기화
        self.ua = NonWindowsUserAgent()

        # 더 정교한 헤더 설정
        self.base_headers = {
            "accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7",
            "accept-encoding": "gzip, deflate, br, zstd",
            "accept-language": "ko-KR,ko;q=0.9,en-US;q=0.8,en;q=0.7",
            "cache-control": "max-age=0",
            "sec-ch-ua": '"Google Chrome";v="131", "Chromium";v="131", "Not_A Brand";v="24"',
            "sec-ch-ua-mobile": "?0",
            "sec-ch-ua-platform": '"macOS"',  # Windows 대신 macOS 사용
            "sec-fetch-dest": "document",
            "sec-fetch-mode": "navigate",
            "sec-fetch-site": "same-origin",
            "sec-fetch-user": "?1",
            "upgrade-insecure-requests": "1",
            "dnt": "1",
        }

        # 쿠키 저장용 세션
        self.session = rq.Session()

        # 헤더에 랜덤 User-Agent 적용
        self.update_headers()

        self.ch = ChromeDriver(self.proxy_rotator)
        self.page_title = None

        # v1.6: URL 매니저 초기화
        self.url_manager = URLManager()

    def get_realistic_headers(self):
        """실제 브라우저와 유사한 헤더 생성 (Windows 제외)"""
        headers = self.base_headers.copy()
        user_agent = self.ua.random
        headers["user-agent"] = user_agent

        # User-Agent에 따라 플랫폼 정보 조정
        if 'iPhone' in user_agent or 'iPad' in user_agent:
            headers["sec-ch-ua-platform"] = '"iOS"'
            headers["sec-ch-ua-mobile"] = "?1" if 'iPhone' in user_agent else "?0"
        elif 'Android' in user_agent:
            headers["sec-ch-ua-platform"] = '"Android"'
            headers["sec-ch-ua-mobile"] = "?1"
        elif 'Macintosh' in user_agent or 'Mac OS X' in user_agent:
            headers["sec-ch-ua-platform"] = '"macOS"'
            headers["sec-ch-ua-mobile"] = "?0"
        elif 'Linux' in user_agent and 'Android' not in user_agent:
            headers["sec-ch-ua-platform"] = '"Linux"'
            headers["sec-ch-ua-mobile"] = "?0"
        else:
            # 기본값은 macOS로 설정 (Windows 방지)
            headers["sec-ch-ua-platform"] = '"macOS"'
            headers["sec-ch-ua-mobile"] = "?0"

        # 랜덤 요소 추가
        if random.choice([True, False]):
            headers["x-requested-with"] = "XMLHttpRequest"

        # 쿠팡 특화 헤더
        headers.update({
            "x-coupang-target-market": "KR",
            "x-coupang-accept-language": "ko-KR",
        })

        return headers

    def update_headers(self):
        """헤더를 새로운 User-Agent로 업데이트"""
        self.headers = self.get_realistic_headers()
        print(f"[DEBUG] 헤더 User-Agent 업데이트: {self.headers['user-agent'][:70]}...")

    def get_session_with_proxy(self):
        """프록시가 적용된 requests 세션 반환"""
        session = rq.Session()
        session.headers.update(self.headers)

        # 더 현실적인 타임아웃 설정
        session.timeout = (10, 30)  # 연결 타임아웃 10초, 읽기 타임아웃 30초

        if self.proxy_rotator and self.proxy_rotator.proxy_list:
            # 성능 우선 랜덤 선택 사용
            proxy = self.proxy_rotator.get_random_proxy_from_working_set()
            if proxy:
                proxy_dict = self.proxy_rotator.get_proxy_dict(proxy)
                if proxy_dict:
                    session.proxies.update(proxy_dict)
                    print(f"[DEBUG] 요청에 프록시 적용: {proxy}")

        return session

    def warm_up_session(self, prod_code):
        """세션을 예열하여 쿠팡 사이트와의 연결을 설정"""
        try:
            print("[INFO] 세션 예열 중...")

            # 메인 페이지 먼저 방문
            main_url = "https://www.coupang.com"
            session = self.get_session_with_proxy()

            # 메인 페이지 방문
            resp = session.get(main_url, timeout=15)
            if resp.status_code == 200:
                print("[DEBUG] 메인 페이지 방문 성공")

                # 쿠키 업데이트
                self.session.cookies.update(resp.cookies)

                # 잠시 대기
                time.sleep(random.uniform(2, 4))

                # 상품 페이지 방문
                product_url = f"https://www.coupang.com/vp/products/{prod_code}"
                resp2 = session.get(product_url, timeout=15)

                if resp2.status_code == 200:
                    print("[DEBUG] 상품 페이지 방문 성공")
                    self.session.cookies.update(resp2.cookies)
                    return True

        except Exception as e:
            print(f"[WARNING] 세션 예열 실패: {e}")

        return False

    def get_product_title(self, product_name: str) -> str:
        """JSON에서 가져온 상품명 사용"""
        print(f"[DEBUG] JSON에서 가져온 상품명 사용: {product_name}")
        return product_name

    def is_timeout_error(self, exception) -> bool:
        """타임아웃 관련 예외인지 확인"""
        return isinstance(exception, (Timeout, ConnectTimeout, ReadTimeout)) or \
            (isinstance(exception, RequestException) and "timeout" in str(exception).lower())

    def handle_consecutive_timeouts(self) -> None:
        """연속 타임아웃 처리"""
        if self.consecutive_timeouts >= self.max_consecutive_timeouts:
            wait_time = random.uniform(self.long_wait_min, self.long_wait_max)
            wait_minutes = wait_time / 60
            print(f"[WARNING] 연속 {self.consecutive_timeouts}회 타임아웃 발생!")
            print(f"[INFO] 서버 안정화를 위해 {wait_minutes:.1f}분 대기합니다...")

            remaining_time = wait_time
            while remaining_time > 0:
                minutes_left = remaining_time / 60
                print(f"[INFO] 남은 대기 시간: {minutes_left:.1f}분")

                sleep_duration = min(30, remaining_time)
                time.sleep(sleep_duration)
                remaining_time -= sleep_duration

            print(f"[INFO] 대기 완료! 크롤링을 재개합니다.")
            self.consecutive_timeouts = 0

    def start(self) -> None:
        """v1.7: 다중 상품 처리를 위한 메인 시작 함수 (JSON 지원)"""
        print("=" * 70)
        print("🛒 쿠팡 리뷰 크롤러 v1.7 (JSON 지원 + 랜덤 프록시)")
        print("=" * 70)

        # JSON 파일 로드
        if not self.url_manager.load_urls_from_json():
            print("[ERROR] JSON 파일을 로드할 수 없습니다.")
            return

        total_products = len(self.url_manager.products)
        print(f"[INFO] 총 {total_products}개 상품을 순차적으로 크롤링합니다.")
        print(f"[INFO] 각 상품당 최대 {self.max_pages}페이지까지 크롤링합니다.")
        print(f"[INFO] 연속 5번 리뷰 없음 감지시 다음 상품으로 진행합니다.")

        # 프록시 사용 정보 출력
        if self.proxy_rotator and self.proxy_rotator.proxy_list:
            available_proxies = self.proxy_rotator.get_available_proxy_count()
            print(f"[INFO] 사용 가능한 프록시: {available_proxies}/{len(self.proxy_rotator.proxy_list)}개")
            print(f"[INFO] 🎲 프록시 랜덤 선택 모드 활성화")

        print("=" * 70)

        # 전체 통계
        total_success_products = 0
        total_failed_products = 0
        overall_start_time = time.time()

        # 상품별 크롤링 실행
        while True:
            product = self.url_manager.get_next_product()
            if not product:
                break

            current_progress, total_progress = self.url_manager.get_current_progress()
            print(f"\n{'=' * 20} 상품 {current_progress}/{total_progress} {'=' * 20}")
            print(f"[INFO] 현재 상품: {product['name']}")
            print(f"[INFO] 상품 URL: {product['url']}")

            try:
                success = self.crawl_single_product(product['url'], product['name'])
                if success:
                    total_success_products += 1
                    print(f"✅ 상품 {current_progress} 크롤링 성공")
                else:
                    total_failed_products += 1
                    print(f"❌ 상품 {current_progress} 크롤링 실패")

            except KeyboardInterrupt:
                print(f"\n[INFO] 사용자에 의해 중단되었습니다.")
                print(f"[INFO] 진행률: {current_progress - 1}/{total_progress} 완료")
                break
            except Exception as e:
                print(f"[ERROR] 상품 크롤링 중 예외 발생: {e}")
                total_failed_products += 1
                continue

            # 상품 간 대기 시간
            if self.url_manager.get_remaining_count() > 0:
                delay = random.uniform(10, 20)  # 상품 간 10-20초 대기
                print(f"[INFO] 다음 상품까지 {delay:.1f}초 대기...")
                time.sleep(delay)

        # 전체 결과 요약
        overall_end_time = time.time()
        total_elapsed = overall_end_time - overall_start_time

        print("\n" + "=" * 70)
        print("📊 전체 크롤링 결과 요약")
        print("=" * 70)
        print(f"총 상품 수: {total_products}개")
        print(f"성공한 상품: {total_success_products}개")
        print(f"실패한 상품: {total_failed_products}개")
        print(f"성공률: {(total_success_products / total_products * 100):.1f}%")
        print(f"총 소요 시간: {total_elapsed / 60:.1f}분")
        print(f"📁 결과 파일들은 'Coupang-reviews' 폴더에서 확인하세요.")
        print("=" * 70)

    def crawl_single_product(self, url: str, product_name: str) -> bool:
        """단일 상품 크롤링"""
        if '#' in url:
            url = url.split('#')[0]
            print(f"[DEBUG] URL fragment 제거: {url}")

        prod_code: str = self.get_product_code(url=url)
        print(f"[DEBUG] 상품 코드: {prod_code}")

        # 세션 예열
        self.warm_up_session(prod_code)

        # 상품별 SaveData 인스턴스 생성
        sd = SaveData()

        try:
            self.title = self.get_product_title(product_name=product_name)
            print(f"[INFO] 상품명: {self.title}")
        except Exception as e:
            print(f"[ERROR] 상품명을 불러오는 도중 오류가 발생했습니다: {e}")
            self.title = "상품명 미확인"

        self.page_title = None  # 페이지 타이틀 초기화
        success_count = 0
        current_page = 1
        consecutive_empty_pages = 0
        max_empty_pages = 5  # v1.6: 연속 빈 페이지 허용 횟수 (5번 연속 리뷰 없음시 다음 상품으로)
        proxy_change_attempts = 0

        product_start_time = time.time()

        while consecutive_empty_pages < max_empty_pages and current_page <= self.max_pages:
            payload = {
                "productId": prod_code,
                "page": current_page,
                "size": 10,
                "sortBy": "DATE_DESC",
                "ratings": "",
                "q": "",
                "viRoleCode": 2,
                "ratingSummary": False,
            }

            result = self.fetch(payload=payload, sd=sd)

            if result:
                success_count += 1
                consecutive_empty_pages = 0
                proxy_change_attempts = 0
            else:
                consecutive_empty_pages += 1
                print(f"[WARNING] 페이지 {current_page}에서 리뷰를 찾을 수 없습니다. ({consecutive_empty_pages}/{max_empty_pages})")

                # 연속 빈 페이지가 2개 이상이고 프록시를 사용 중이라면 프록시 상태 체크
                if (consecutive_empty_pages >= 2 and
                        self.proxy_rotator and
                        self.proxy_rotator.current_proxy and
                        proxy_change_attempts < 3):

                    available_proxies = self.proxy_rotator.get_available_proxy_count()
                    if available_proxies > 1:
                        print(f"[INFO] 연속 실패로 인한 프록시 교체 시도 ({proxy_change_attempts + 1}/3)")
                        self.proxy_rotator.mark_proxy_failed(self.proxy_rotator.current_proxy)
                        proxy_change_attempts += 1
                        print(f"[INFO] 페이지 {current_page} 다른 프록시로 재시도...")
                        continue

            current_page += 1

            if result and consecutive_empty_pages == 0:
                short_delay = random.uniform(1.0, 3.0)
                time.sleep(short_delay)

        product_end_time = time.time()
        product_elapsed = product_end_time - product_start_time

        # 상품별 결과 출력
        print(f"\n[PRODUCT SUMMARY] 상품 '{self.title}' 크롤링 완료")
        print(f"[INFO] 성공 페이지: {success_count}개 (총 {current_page - 1}페이지 시도)")
        print(f"[INFO] 소요 시간: {product_elapsed / 60:.1f}분")

        if consecutive_empty_pages >= max_empty_pages:
            print(f"[INFO] 연속 {max_empty_pages}번 빈 페이지로 인해 다음 상품으로 진행")
        elif current_page > self.max_pages:
            print(f"[INFO] 최대 페이지 수({self.max_pages})에 도달하여 완료")

        return success_count > 0

    def fetch(self, payload: dict, sd) -> bool:
        now_page: int = payload["page"]
        print(f"\n[INFO] Start crawling page {now_page} ...")
        attempt: int = 0
        proxy_attempts: int = 0
        max_proxy_attempts: int = min(10, len(self.proxy_rotator.proxy_list) if self.proxy_rotator else 0)

        while attempt < self.retries:
            try:
                # 매 요청마다 새로운 User-Agent 사용
                if attempt > 0:
                    self.update_headers()

                session = self.get_session_with_proxy()
                session.cookies.update(self.session.cookies)
                session.headers.update({
                    "Referer": f"https://www.coupang.com/vp/products/{payload['productId']}"
                })

                resp = session.get(
                    url=self.base_review_url,
                    params=payload,
                    timeout=(15, 30),
                )

                self.consecutive_timeouts = 0

                if resp.status_code == 403:
                    print(f"[ERROR] HTTP 403 응답 - 프록시가 차단됨")
                    if self.proxy_rotator and self.proxy_rotator.current_proxy:
                        self.proxy_rotator.mark_proxy_failed(self.proxy_rotator.current_proxy)
                    attempt += 1
                    continue
                elif resp.status_code != 200:
                    print(f"[ERROR] HTTP {resp.status_code} 응답")
                    attempt += 1
                    continue

                html = resp.text
                soup = bs(html, "html.parser")

                if self.page_title is None:
                    first_review = soup.select_one("article.sdp-review__article__list")
                    if first_review:
                        title_elem = first_review.select_one("div.sdp-review__article__list__info__product-info__name")
                        self.page_title = title_elem.text.strip() if title_elem else self.title
                    else:
                        self.page_title = self.title

                articles = soup.select("article.sdp-review__article__list")
                article_length = len(articles)

                if article_length == 0:
                    print(f"[WARNING] 페이지 {now_page}에서 리뷰를 찾을 수 없습니다.")

                    # 프록시 사용 중이라면 다른 프록시로 재시도
                    if self.proxy_rotator and self.proxy_rotator.current_proxy and proxy_attempts < max_proxy_attempts:
                        print(f"[INFO] 프록시 차단 가능성으로 다른 프록시로 재시도 ({proxy_attempts + 1}/{max_proxy_attempts})")
                        self.proxy_rotator.mark_proxy_failed(self.proxy_rotator.current_proxy)
                        proxy_attempts += 1
                        attempt += 1
                        retry_delay = random.uniform(1.0, 3.0)
                        print(f"[DEBUG] {retry_delay:.1f}초 후 다른 프록시로 재시도...")
                        time.sleep(retry_delay)
                        continue

                    # 차단 감지 및 추가 처리
                    if now_page == 1:
                        print("[DEBUG] 첫 페이지 HTML 구조 확인:")
                        print(f"  - 전체 길이: {len(html)} 문자")
                        print(f"  - 'review' 포함 횟수: {html.lower().count('review')}")
                        print(f"  - 'article' 포함 횟수: {html.lower().count('article')}")

                        blocked_indicators = [
                            "access denied", "blocked", "forbidden",
                            "captcha", "robot", "bot", "security", "verification"
                        ]

                        html_lower = html.lower()
                        is_blocked = False
                        for indicator in blocked_indicators:
                            if indicator in html_lower:
                                print(f"[WARNING] 차단 감지: '{indicator}' 발견")
                                is_blocked = True
                                break

                        if is_blocked and attempt < self.retries - 2:
                            print("[INFO] 차단 감지로 인한 추가 재시도...")
                            attempt += 1
                            long_delay = random.uniform(5.0, 10.0)
                            print(f"[DEBUG] {long_delay:.1f}초 대기 후 재시도...")
                            time.sleep(long_delay)
                            continue

                    return False

                print(f"[SUCCESS] 페이지 {now_page}에서 {article_length}개 리뷰 발견")

                # 리뷰 데이터 처리
                for idx in range(article_length):
                    dict_data: dict[str, str | int] = dict()

                    review_date_elem = articles[idx].select_one(
                        "div.sdp-review__article__list__info__product-info__reg-date"
                    )
                    review_date = review_date_elem.text.strip() if review_date_elem else "-"

                    user_name_elem = articles[idx].select_one(
                        "span.sdp-review__article__list__info__user__name"
                    )
                    user_name = user_name_elem.text.strip() if user_name_elem else "-"

                    rating_elem = articles[idx].select_one(
                        "div.sdp-review__article__list__info__product-info__star-orange"
                    )
                    if rating_elem and rating_elem.get("data-rating"):
                        try:
                            rating = int(rating_elem.get("data-rating"))
                        except (ValueError, TypeError):
                            rating = 0
                    else:
                        rating = 0

                    prod_name_elem = articles[idx].select_one(
                        "div.sdp-review__article__list__info__product-info__name"
                    )
                    prod_name = prod_name_elem.text.strip() if prod_name_elem else "-"

                    headline_elem = articles[idx].select_one(
                        "div.sdp-review__article__list__headline"
                    )
                    headline = headline_elem.text.strip() if headline_elem else ""

                    review_content_elem = articles[idx].select_one(
                        "div.sdp-review__article__list__review__content.js_reviewArticleContent"
                    )
                    if review_content_elem:
                        review_content = re.sub("[\n\t]", "", review_content_elem.text.strip())
                    else:
                        review_content_elem = articles[idx].select_one(
                            "div.sdp-review__article__list__review > div"
                        )
                        if review_content_elem:
                            review_content = re.sub("[\n\t]", "", review_content_elem.text.strip())
                        else:
                            review_content = ""

                    helpful_count_elem = articles[idx].select_one("span.js_reviewArticleHelpfulCount")
                    helpful_count = helpful_count_elem.text.strip() if helpful_count_elem else "0"

                    review_images = articles[idx].select("div.sdp-review__article__list__attachment__list img")
                    image_count = len(review_images)

                    dict_data["title"] = self.page_title
                    dict_data["prod_name"] = prod_name
                    dict_data["review_date"] = review_date
                    dict_data["user_name"] = user_name
                    dict_data["rating"] = rating
                    dict_data["headline"] = headline
                    dict_data["review_content"] = review_content
                    dict_data["helpful_count"] = helpful_count
                    dict_data["image_count"] = image_count

                    sd.save(datas=dict_data)
                    print(f"[SUCCESS] 리뷰 저장 완료: {user_name} - {rating}점")

                page_delay = random.uniform(self.page_delay_min, self.page_delay_max)
                print(f"[DEBUG] 다음 페이지까지 {page_delay:.1f}초 대기...")
                time.sleep(page_delay)
                return True

            except RequestException as e:
                attempt += 1

                error_str = str(e).lower()
                is_proxy_error = any(keyword in error_str for keyword in [
                    "403", "proxy", "connection", "timeout", "refused", "unreachable"
                ])

                if is_proxy_error and self.proxy_rotator and self.proxy_rotator.current_proxy:
                    self.proxy_rotator.mark_proxy_failed(self.proxy_rotator.current_proxy)
                    print("[INFO] 프록시 오류로 인한 다른 프록시로 재시도합니다.")

                    available_proxies = self.proxy_rotator.get_available_proxy_count()
                    if available_proxies > 0:
                        print(f"[INFO] 남은 사용 가능 프록시: {available_proxies}개")
                    else:
                        print("[WARNING] 사용 가능한 프록시가 없습니다.")

                if self.is_timeout_error(e):
                    self.consecutive_timeouts += 1
                    print(f"[ERROR] 타임아웃 발생 (연속 {self.consecutive_timeouts}회): {e}")

                    if self.consecutive_timeouts >= self.max_consecutive_timeouts:
                        self.handle_consecutive_timeouts()
                else:
                    self.consecutive_timeouts = 0
                    print(f"[ERROR] 네트워크 오류: {e}")

                print(f"[ERROR] Attempt {attempt}/{self.retries} failed")
                if attempt < self.retries:
                    retry_delay = random.uniform(self.delay_min, self.delay_max)
                    print(f"[DEBUG] {retry_delay:.1f}초 후 재시도...")
                    time.sleep(retry_delay)
                else:
                    print(f"[ERROR] 최대 요청 횟수 초과! 페이지 {now_page} 크롤링 실패.")
                    return False
            except Exception as e:
                print(f"[ERROR] 예상치 못한 오류 발생: {e}")
                self.consecutive_timeouts = 0
                return False

        return False

    @staticmethod
    def clear_console() -> None:
        command: str = "clear"
        if os.name in ("nt", "dos"):
            command = "cls"
        try:
            if os.environ.get('TERM') is None:
                os.environ['TERM'] = 'xterm'
            os.system(command=command)
        except:
            pass


class SaveData:
    def __init__(self) -> None:
        self.wb: Workbook = Workbook()
        self.ws = self.wb.active
        self.ws.append([
            "상품명", "구매상품명", "작성일자", "구매자명", "평점",
            "헤드라인", "리뷰내용", "도움수", "이미지수"
        ])
        self.row: int = 2
        self.dir_name: str = "data/Coupang-reviews-homeplanet"
        self.create_directory()

    def create_directory(self) -> None:
        if not os.path.exists(self.dir_name):
            os.makedirs(self.dir_name)
            print(f"[INFO] 디렉토리 생성: {self.dir_name}")

    def save(self, datas: dict[str, str | int]) -> None:
        try:
            safe_title = re.sub(r'[<>:"/\\|?*]', '_', datas["title"])
            file_name: str = os.path.join(self.dir_name, safe_title + ".xlsx")

            self.ws[f"A{self.row}"] = datas["title"]
            self.ws[f"B{self.row}"] = datas["prod_name"]
            self.ws[f"C{self.row}"] = datas["review_date"]
            self.ws[f"D{self.row}"] = datas["user_name"]
            self.ws[f"E{self.row}"] = datas["rating"]
            self.ws[f"F{self.row}"] = datas["headline"]
            self.ws[f"G{self.row}"] = datas["review_content"]
            self.ws[f"H{self.row}"] = datas["helpful_count"]
            self.ws[f"I{self.row}"] = datas["image_count"]

            self.row += 1
            self.wb.save(filename=file_name)

        except Exception as e:
            print(f"[ERROR] 데이터 저장 중 오류 발생: {e}")

    def __del__(self) -> None:
        try:
            if hasattr(self, 'wb'):
                self.wb.close()
        except:
            pass


def load_proxy_list_from_file(file_path="env/proxy_list.txt"):
    """txt 파일에서 프록시 목록 로드"""
    try:
        if not os.path.exists(file_path):
            create_sample_proxy_file(file_path)
            return []

        with open(file_path, 'r', encoding='utf-8') as f:
            lines = f.readlines()

        proxy_list = []

        for line_num, line in enumerate(lines, 1):
            line = line.strip()

            # 빈 줄이나 주석(#으로 시작) 건너뛰기
            if not line or line.startswith('#'):
                continue

            # 프록시 형식 검증 (ip:port:username:password)
            parts = line.split(':')
            if len(parts) == 4:
                ip, port, username, password = parts
                # 기본적인 IP와 포트 검증
                if is_valid_proxy_format(ip, port):
                    proxy_list.append(line)
                    print(f"[INFO] 프록시 로드: {ip}:{port}")
                else:
                    print(f"[WARNING] 잘못된 프록시 형식 (라인 {line_num}): {line}")
            else:
                print(f"[WARNING] 잘못된 형식 (라인 {line_num}): {line}")
                print("         올바른 형식: ip:port:username:password")

        if proxy_list:
            print(f"[SUCCESS] {len(proxy_list)}개의 유효한 프록시를 로드했습니다.")
            return proxy_list
        else:
            print("[ERROR] 유효한 프록시가 없습니다.")
            return []

    except Exception as e:
        print(f"[ERROR] 프록시 파일 읽기 실패: {e}")
        return []


def create_sample_proxy_file(file_path="proxy_list.txt"):
    """샘플 프록시 파일 생성"""
    sample_content = """# 프록시 목록 파일
# 형식: ip:port:username:password
# 한 줄에 하나씩 입력하세요
# '#'으로 시작하는 줄은 주석으로 처리됩니다

# 샘플 프록시 (실제 프록시로 교체하세요)
173.214.177.18:5709:daxvymvx:kn518nmfd34a
198.23.214.119:6386:daxvymvx:kn518nmfd34a
50.114.98.49:5533:daxvymvx:kn518nmfd34a

# 추가 프록시들을 여기에 입력하세요
# 192.168.1.1:8080:user:pass
# 10.0.0.1:3128:admin:password
"""

    try:
        with open(file_path, 'w', encoding='utf-8') as f:
            f.write(sample_content)
        print(f"[INFO] 샘플 프록시 파일이 생성되었습니다: {file_path}")
        print("[INFO] 파일을 편집하여 실제 프록시 정보를 입력한 후 다시 실행하세요.")
    except Exception as e:
        print(f"[ERROR] 샘플 파일 생성 실패: {e}")


def is_valid_proxy_format(ip, port):
    """기본적인 IP와 포트 형식 검증"""
    try:
        # IP 주소 형식 검증 (간단한 검증)
        ip_parts = ip.split('.')
        if len(ip_parts) != 4:
            return False

        for part in ip_parts:
            if not part.isdigit() or not (0 <= int(part) <= 255):
                return False

        # 포트 번호 검증
        if not port.isdigit() or not (1 <= int(port) <= 65535):
            return False

        return True
    except:
        return False


def test_proxy(proxy_string):
    """프록시 연결 테스트"""
    try:
        parts = proxy_string.split(':')
        if len(parts) == 4:
            ip, port, username, password = parts
            proxy_dict = {
                'http': f'http://{username}:{password}@{ip}:{port}',
                'https': f'http://{username}:{password}@{ip}:{port}'
            }
        else:
            return False

        response = rq.get('http://httpbin.org/ip', proxies=proxy_dict, timeout=10)
        if response.status_code == 200:
            return True
    except:
        pass
    return False


def get_proxy_list():
    """프록시 목록 반환 (파일에서 로드)"""
    proxy_file_path = "env/proxy_list.txt"

    print("=" * 70)
    print("🔗 프록시 설정")
    print("=" * 70)

    # 프록시 파일에서 로드
    proxy_list = load_proxy_list_from_file(proxy_file_path)

    if not proxy_list:
        print(f"\n[WARNING] {proxy_file_path} 파일에 유효한 프록시가 없습니다.")

        # 프록시 없이 실행할지 확인
        run_without_proxy = input("프록시 없이 실행하시겠습니까? (Y/n): ").lower().strip()
        if run_without_proxy != 'n':
            print("[INFO] 프록시 없이 실행합니다.")
            return None
        else:
            print("[INFO] 프로그램을 종료합니다.")
            print(f"[INFO] {proxy_file_path} 파일을 편집하여 프록시를 추가한 후 다시 실행하세요.")
            exit(0)

    print(f"\n[INFO] {len(proxy_list)}개의 프록시가 준비되었습니다.")
    print()

    # 프록시 사용 여부 확인
    use_proxy = input("프록시를 사용하시겠습니까? (Y/n): ").lower().strip()

    if use_proxy == 'n':
        print("[INFO] 프록시 없이 실행합니다.")
        return None
    else:
        print("[INFO] 프록시를 사용하여 실행합니다.")

        # 프록시 테스트 여부 확인
        test_proxies = input("프록시 연결을 테스트하시겠습니까? (y/N): ").lower().strip()

        if test_proxies == 'y':
            print("\n[INFO] 프록시 연결 테스트 중...")
            working_proxies = []

            for i, proxy in enumerate(proxy_list, 1):
                print(f"[TEST] {i}/{len(proxy_list)} - {proxy.split(':')[0]}:{proxy.split(':')[1]} 테스트 중...", end='')
                if test_proxy(proxy):
                    print(" 성공")
                    working_proxies.append(proxy)
                else:
                    print(" 실패")

            if working_proxies:
                print(f"\n[SUCCESS] {len(working_proxies)}/{len(proxy_list)}개 프록시가 정상 작동합니다.")
                print(f"[INFO] 작동하는 프록시만 사용하여 크롤링을 시작합니다.")
                return working_proxies
            else:
                print("\n[ERROR] 작동하는 프록시가 없습니다.")
                fallback = input("프록시 없이 실행하시겠습니까? (Y/n): ").lower().strip()
                if fallback != 'n':
                    print("[INFO] 프록시 없이 실행합니다.")
                    return None
                else:
                    print("[INFO] 프로그램을 종료합니다.")
                    exit(0)
        else:
            print("[INFO] 테스트 없이 모든 프록시를 사용합니다.")
            print("[INFO] 실행 중 자동으로 작동하지 않는 프록시를 제외합니다.")
            return proxy_list


if __name__ == "__main__":
    try:
        # 프록시 목록 가져오기
        proxy_list = get_proxy_list()

        # 크롤러 시작
        coupang = Coupang(proxy_list=proxy_list)
        coupang.start()

        print("\n" + "=" * 70)
        print("모든 상품 크롤링이 완료되었습니다!")
        print("결과 파일들은 'Coupang-reviews' 폴더에서 확인하세요.")
        print("=" * 70)

    except KeyboardInterrupt:
        print("\n[INFO] 사용자에 의해 중단되었습니다.")
    except Exception as e:
        print(f"\n[ERROR] 프로그램 실행 중 오류 발생: {e}")
        import traceback

        traceback.print_exc()
    finally:
        print("\n프로그램을 종료합니다.")